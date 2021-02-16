#! python3

'''
1. import excel sheet
2. loop through columns to find text
3. for each column, write the code on each row to a specific text doc
4. save text docs
'''

import openpyxl

wb = openpyxl.load_workbook('testDoc.xlsx')
sheet = wb.active

for col in range(sheet.max_column):
    newDoc = open('writtendoc%s.txt' % str(col+1), 'w')
    for row in range(sheet.max_row):
        try:
            newDoc.write(sheet.cell(row=row+1, column=col+1).value)
        except TypeError:
            break
    newDoc.close()
