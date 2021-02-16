#! python3

'''
1. create an empty dictionary
2. read values from each cell and assign it to the dictionary based on its location
3. copy results onto the excel sheet
'''

import openpyxl

file = 'example.xlsx'
newFile = 'testdoc.xlsx'

input_wb = openpyxl.load_workbook(file)
input_sheet = input_wb.active
output_wb = openpyxl.Workbook()
output_sheet = output_wb.active

maxrow = input_sheet.max_row
maxcol = input_sheet.max_column

for x in range(1, maxrow + 1):
    for y in range(1, maxcol + 1):
        #output_sheet.cell(row=x, column=y).value = input_sheet.cell(row=y, column=x).value
        output_sheet.cell(row=y, column=x).value = input_sheet.cell(row=x, column=y).value

input_wb.save(file)
output_wb.save(newFile)
